"""
Парсер Excel-файла с программой тренировок.
Читает листы НЕДЕЛЯ 1-4 и строит структуру тренировок.
"""

import re
import openpyxl
from dataclasses import dataclass, field
from typing import Optional


@dataclass
class ExerciseSet:
    reps: str          # e.g. "10" or "10/10/10" (по подходам)
    weight: str        # e.g. "40" или "Б/в" или "Разм.:35 | Раб.:41"
    rest_seconds: int


@dataclass
class Exercise:
    name: str
    sets: int
    reps_per_set: list[str]   # список повторений на каждый подход
    weight: str               # сырое значение веса
    weight_kg: Optional[float]  # рабочий вес в кг (если удалось распарсить)
    rest_seconds: int
    comment: str


@dataclass
class Workout:
    week: int
    number: int           # номер тренировки (1–12)
    name: str             # e.g. "Грудь + Плечи + Пресс"
    sheet_name: str
    exercises: list[Exercise] = field(default_factory=list)


def parse_weight(raw: str) -> Optional[float]:
    """Извлекает рабочий вес из строки. Возвращает None если не числовой."""
    if not raw or str(raw).strip() in ('Б/в', '', 'б/в'):
        return None
    raw = str(raw).strip()
    # "Разм.:35 | Раб.:41" → 41
    m = re.search(r'Раб\.?:?\s*([\d.]+)', raw, re.IGNORECASE)
    if m:
        return float(m.group(1))
    # "40–42.5" → берём первое число
    m = re.search(r'([\d.]+)', raw)
    if m:
        return float(m.group(1))
    return None


def parse_reps(reps_str: str, sets: int) -> list[str]:
    """Разбивает строку повторений на список по подходам."""
    if not reps_str:
        return ['?'] * sets
    parts = str(reps_str).split('/')
    # Если меньше частей чем подходов — дублируем последнюю
    while len(parts) < sets:
        parts.append(parts[-1])
    return parts[:sets]


def parse_rest(rest_val) -> int:
    """Парсит отдых в секунды."""
    if rest_val is None:
        return 90
    try:
        return int(str(rest_val).strip())
    except (ValueError, AttributeError):
        return 90


def is_header_row(row) -> bool:
    """Проверяет, является ли строка заголовком таблицы."""
    return row[0] == 'Упражнение'


def is_workout_title(row) -> Optional[tuple[int, str]]:
    """
    Если строка — заголовок тренировки, возвращает (номер, название).
    Например: 'Тренировка 3 — Спина + Трицепс + Пресс' → (3, 'Спина + Трицепс + Пресс')
    """
    if row[0] and all(row[i] is None for i in range(1, len(row))):
        m = re.match(r'Тренировка\s+(\d+)\s*[—–-]\s*(.+)', str(row[0]))
        if m:
            num = int(m.group(1))
            name = re.sub(r'\s*[↑↓]\s*.*$', '', m.group(2)).strip()  # убираем "↑ нагрузка" и т.п.
            name = re.sub(r'\s*\(тест\)', '', name).strip()
            return num, name
    return None


def parse_exercise_row(row) -> Optional[Exercise]:
    """Парсит строку упражнения."""
    if not row[0] or row[0] == 'Упражнение':
        return None

    name = str(row[0]).strip()
    # Пропускаем строки-заголовки недели
    if re.match(r'НЕДЕЛЯ\s+\d', name):
        return None

    try:
        sets = int(str(row[1]).strip()) if row[1] else 3
    except (ValueError, AttributeError):
        sets = 3

    reps_str = str(row[2]) if row[2] else '10'
    reps_list = parse_reps(reps_str, sets)

    weight_raw = str(row[3]) if row[3] is not None else 'Б/в'
    weight_kg = parse_weight(weight_raw)
    rest_sec = parse_rest(row[4])
    comment = str(row[5]) if row[5] else ''

    return Exercise(
        name=name,
        sets=sets,
        reps_per_set=reps_list,
        weight=weight_raw,
        weight_kg=weight_kg,
        rest_seconds=rest_sec,
        comment=comment,
    )


def parse_workouts(filepath: str) -> list[Workout]:
    """Читает Excel и возвращает список всех тренировок."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    workouts: list[Workout] = []

    week_sheets = [s for s in wb.sheetnames if re.match(r'НЕДЕЛЯ\s+\d', s)]

    for sheet_name in week_sheets:
        m = re.search(r'\d+', sheet_name)
        week_num = int(m.group()) if m else 0
        ws = wb[sheet_name]

        current_workout: Optional[Workout] = None

        for row in ws.iter_rows(values_only=True):
            if all(v is None for v in row):
                continue

            # Проверяем заголовок тренировки
            title = is_workout_title(row)
            if title:
                workout_num, workout_name = title
                current_workout = Workout(
                    week=week_num,
                    number=workout_num,
                    name=workout_name,
                    sheet_name=sheet_name,
                )
                workouts.append(current_workout)
                continue

            # Пропускаем строку-заголовок таблицы и заголовок недели
            if is_header_row(row):
                continue
            if row[0] and re.match(r'НЕДЕЛЯ\s+\d', str(row[0])):
                continue

            # Парсим упражнение
            if current_workout:
                ex = parse_exercise_row(row)
                if ex:
                    current_workout.exercises.append(ex)

    workouts.sort(key=lambda w: w.number)
    return workouts


def print_mapping(workouts: list[Workout]):
    """Выводит маппинг для проверки пользователем."""
    print("\n" + "="*70)
    print("МАППИНГ ТРЕНИРОВОК (для проверки перед загрузкой)")
    print("="*70)
    for w in workouts:
        print(f"\n[Тренировка {w.number}] Неделя {w.week} — {w.name}")
        print(f"  Упражнений: {len(w.exercises)}")
        for i, ex in enumerate(w.exercises, 1):
            weight_str = f"{ex.weight_kg} кг" if ex.weight_kg else ex.weight
            print(f"  {i:2}. {ex.name}")
            print(f"      Подходы: {ex.sets}  |  Повторения: {'/'.join(ex.reps_per_set)}  |  Вес: {weight_str}  |  Отдых: {ex.rest_seconds}с")
    print("\n" + "="*70)
    print(f"Итого: {len(workouts)} тренировок")
    print("="*70)


if __name__ == '__main__':
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else r'C:\Users\User\Downloads\Программа_тренировок_апрель_2026.xlsx'
    workouts = parse_workouts(path)
    print_mapping(workouts)
